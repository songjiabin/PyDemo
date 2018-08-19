# xlsx xls

from openpyxl import load_workbook


def readXlsxFile(path):
    # 打开表格
    file = load_workbook(filename=path)
    # 所以表格的名称 file.get_sheet_names()
    print(file.get_sheet_names())

    # 拿出一个表格 根据名字
    sheet = file.get_sheet_by_name(file.get_sheet_names()[0])
    # 一共有多少行
    print(sheet.max_row)
    # 一共有多少列
    print(sheet.max_column)
    # 这个表格的表名
    print(sheet.title)

    # 读取数据 一张表
    for lineNum in range(1, sheet.max_row + 1):
        # 在每行中处理每列
        lineList = []
        for columnNum in range(1, sheet.max_column + 1):
            # 那数据
            value = sheet.cell(row=lineNum, column=columnNum).value

            lineList.append(value)
        pass
        print(lineList)

    pass


def main():
    path = r"E:\py_project_path\23excel文件处理\res\b.xlsx"
    readXlsxFile(path)
    pass


if __name__ == '__main__':
    main()
