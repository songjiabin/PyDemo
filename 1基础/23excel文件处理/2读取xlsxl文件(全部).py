# xlsx xls

from openpyxl import load_workbook


def readXlsxFile(path):
    # 打开表格

    file = load_workbook(filename=path)
    # 所以表格的名称 file.get_sheet_names()
    print(file.get_sheet_names())

    allExcelNum = file.get_sheet_names()

    # 总的大字典
    dic = {}

    for sheetName in allExcelNum:

        sheetInfo = []

        sheet = file.get_sheet_by_name(sheetName)

        # 读取数据 一张表
        for lineNum in range(1, sheet.max_row + 1):
            # 在每行中处理每列
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                # 那数据
                value = sheet.cell(row=lineNum, column=columnNum).value

                lineList.append(value)
            sheetInfo.append(lineList)
        dic[sheetName] = sheetInfo
    return  dic
    pass


def main():
    path = r"E:\py_project_path\23excel文件处理\res\b.xlsx"
    result =readXlsxFile(path)
    print(result)
    pass


if __name__ == '__main__':
    main()
